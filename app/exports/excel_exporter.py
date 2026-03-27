from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.exports.institutional_layout import (
    BLACK,
    EXCEL_COLUMN_WIDTHS,
    INSTITUTIONAL_GRAY,
    INSTITUTIONAL_GREEN,
    INSTITUTIONAL_GREEN_LIGHT,
    TABLE_HEADERS,
    TABLE_GRID,
    WHITE,
    get_logo_path,
)
from app.exports.report_builder import ReportBuilder


class ExcelExporter:
    @staticmethod
    def export_group_schedule(group_id: int) -> BytesIO:
        context = ReportBuilder.build_group_report_context(group_id)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Grupo {context['group_number']}"
        sheet.sheet_view.showGridLines = False

        for column, width in EXCEL_COLUMN_WIDTHS.items():
            sheet.column_dimensions[column].width = width

        ExcelExporter._render_header(sheet, context)
        ExcelExporter._render_table(sheet, context)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _render_header(sheet, context: dict) -> None:
        green_fill = PatternFill("solid", fgColor=INSTITUTIONAL_GREEN)
        light_fill = PatternFill("solid", fgColor=INSTITUTIONAL_GREEN_LIGHT)
        white_font = Font(color=WHITE, bold=True, size=13)
        title_font = Font(color=INSTITUTIONAL_GRAY, bold=True, size=12)
        normal_font = Font(color=BLACK, size=10)
        small_font = Font(color=INSTITUTIONAL_GRAY, size=9)
        thin_border = Border(
            left=Side(style="thin", color=TABLE_GRID),
            right=Side(style="thin", color=TABLE_GRID),
            top=Side(style="thin", color=TABLE_GRID),
            bottom=Side(style="thin", color=TABLE_GRID),
        )

        sheet.merge_cells("A1:B1")
        sheet["A1"] = context["generated_time"]
        sheet["A1"].font = normal_font

        sheet.merge_cells("M1:N1")
        sheet["M1"] = context["generated_date"]
        sheet["M1"].alignment = Alignment(horizontal="right")
        sheet["M1"].font = normal_font

        sheet.merge_cells("A2:B2")
        sheet["A2"] = context["report_code"]
        sheet["A2"].font = small_font

        sheet.merge_cells("M2:N2")
        sheet["M2"] = "PAG. 1"
        sheet["M2"].alignment = Alignment(horizontal="right")
        sheet["M2"].font = normal_font

        sheet.merge_cells("C1:D4")
        logo_path = get_logo_path()
        if logo_path is not None:
            image = ExcelImage(str(logo_path))
            image.width = 70
            image.height = 78
            sheet.add_image(image, "C1")
        else:
            sheet["C1"] = "LOGO"
            sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
            sheet["C1"].font = Font(bold=True, color=INSTITUTIONAL_GREEN)
            for row in sheet["C1:D4"]:
                for cell in row:
                    cell.border = thin_border
                    cell.fill = light_fill

        sheet.merge_cells("E1:K1")
        sheet["E1"] = context["institution_name"]
        sheet["E1"].font = white_font
        sheet["E1"].alignment = Alignment(horizontal="center")
        sheet["E1"].fill = green_fill

        sheet.merge_cells("E2:K2")
        sheet["E2"] = context["coordination_name"]
        sheet["E2"].font = title_font
        sheet["E2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("E3:K3")
        sheet["E3"] = context["report_title"]
        sheet["E3"].font = title_font
        sheet["E3"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("E4:K4")
        sheet["E4"] = f"Periodo: {context['period']}"
        sheet["E4"].font = small_font
        sheet["E4"].alignment = Alignment(horizontal="center")

        sheet["A5"] = "UNIDAD ACADÉMICA:"
        sheet["B5"] = context["unit_code"]
        sheet.merge_cells("C5:F5")
        sheet["C5"] = context["unit_name"]
        sheet["G5"] = "GRUPO:"
        sheet["N5"] = str(context["group_number"])

        sheet["A6"] = "CARRERA:"
        sheet["B6"] = context["career_code"]
        sheet.merge_cells("C6:F6")
        sheet["C6"] = context["career_name"]

        sheet["A7"] = "PLAN DE ESTUDIOS:"
        sheet["B7"] = context["plan_key"]

        for cell_ref in ["A5", "A6", "A7", "G5"]:
            sheet[cell_ref].font = Font(bold=True, size=10)
        for cell_ref in ["B5", "B6", "B7", "C5", "C6", "N5"]:
            sheet[cell_ref].font = normal_font

    @staticmethod
    def _render_table(sheet, context: dict) -> None:
        header_fill = PatternFill("solid", fgColor=INSTITUTIONAL_GREEN)
        header_font = Font(color=WHITE, bold=True, size=9)
        thin_side = Side(style="thin", color=TABLE_GRID)
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        table_row = 9
        for index, header in enumerate(TABLE_HEADERS, start=1):
            cell = sheet.cell(row=table_row, column=index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_center
            cell.border = border

        current_row = table_row + 1
        row_fields = [
            "clave_control",
            "descripcion_maestro",
            "edificio",
            "salon",
            "cap_asg",
            "tpo_sgp",
            "lunes",
            "martes",
            "miercoles",
            "jueves",
            "viernes",
            "sabado",
            "domingo",
            "es",
        ]

        for report_row in context["rows"]:
            sheet.row_dimensions[current_row].height = 36
            for index, field_name in enumerate(row_fields, start=1):
                cell = sheet.cell(row=current_row, column=index, value=report_row[field_name])
                cell.border = border
                cell.alignment = wrap_left if index in {1, 2} else wrap_center
                cell.font = Font(size=9, bold=index in {1, 2})
            current_row += 1
